"""Quotes export endpoints — CSV and XLSX downloads.

Both endpoints share the extraction layer in ``export_core.py``.
Column headers are passed from the frontend via ``Accept-Language`` or
a ``lang`` query parameter; the server returns English headers by default.
"""

from __future__ import annotations

import csv
import io
import logging

from fastapi import APIRouter, HTTPException, Query, Request
from starlette.responses import Response

from bristlenose.server.export_core import (
    ExportableQuote,
    csv_safe,
    excel_sheet_name,
    extract_quotes_for_export,
)
from bristlenose.server.models import Project
from bristlenose.utils.text import safe_filename

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")

# ---------------------------------------------------------------------------
# Column headers (English defaults — frontend can pass translated headers)
# ---------------------------------------------------------------------------

#: The 11 export columns in order, matching ExportableQuote field order.
_DEFAULT_HEADERS = [
    "Quote",
    "Participant code",
    "Participant name",
    "Section",
    "Theme",
    "Sentiment",
    "Tags",
    "Starred",
    "Timecode",
    "Session",
    "Source file",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _get_db(request: Request):
    """Get the database session from app state."""
    return request.app.state.db_factory()


def _check_project(db, project_id: int) -> Project:
    """Return the project or raise 404."""
    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


def _parse_quote_ids(quote_ids: str | None) -> list[str] | None:
    """Parse comma-separated DOM IDs, or return None for 'all'."""
    if not quote_ids:
        return None
    ids = [qid.strip() for qid in quote_ids.split(",") if qid.strip()]
    return ids or None


def _parse_headers(headers: str | None) -> list[str]:
    """Parse comma-separated translated headers, or return defaults."""
    if not headers:
        return list(_DEFAULT_HEADERS)
    parts = [h.strip() for h in headers.split(",") if h.strip()]
    if len(parts) == len(_DEFAULT_HEADERS):
        return parts
    return list(_DEFAULT_HEADERS)


def _quote_to_row(q: ExportableQuote) -> list[str]:
    """Convert an ExportableQuote to a list of string cell values."""
    return [
        q.text,
        q.participant_code,
        q.participant_name,
        q.section,
        q.theme,
        q.sentiment,
        q.tags,
        "Yes" if q.starred else "",
        q.timecode,
        q.session,
        q.source_file,
    ]


# ---------------------------------------------------------------------------
# CSV endpoint
# ---------------------------------------------------------------------------


@router.get("/projects/{project_id}/export/quotes.csv")
async def export_quotes_csv(
    request: Request,
    project_id: int,
    quote_ids: str | None = Query(None, description="Comma-separated DOM IDs"),
    anonymise: bool = Query(False),
    headers: str | None = Query(None, alias="col_headers",
                                description="Comma-separated translated column headers"),
):
    """Export quotes as CSV with UTF-8 BOM."""
    db = _get_db(request)
    try:
        project = _check_project(db, project_id)
        ids = _parse_quote_ids(quote_ids)
        quotes = extract_quotes_for_export(db, project_id, ids, anonymise=anonymise)

        if not quotes:
            raise HTTPException(status_code=404, detail="No quotes match the filter")

        col_headers = _parse_headers(headers)

        # Build CSV in memory with UTF-8 BOM
        buf = io.StringIO()
        buf.write("\ufeff")  # UTF-8 BOM for Excel on Windows
        writer = csv.writer(buf)
        writer.writerow(col_headers)
        for q in quotes:
            writer.writerow(csv_safe(v) for v in _quote_to_row(q))

        filename = f"{safe_filename(project.name)}-quotes.csv"
        return Response(
            content=buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )
    finally:
        db.close()


# ---------------------------------------------------------------------------
# XLSX endpoint
# ---------------------------------------------------------------------------


@router.get("/projects/{project_id}/export/quotes.xlsx")
async def export_quotes_xlsx(
    request: Request,
    project_id: int,
    quote_ids: str | None = Query(None, description="Comma-separated DOM IDs"),
    anonymise: bool = Query(False),
    headers: str | None = Query(None, alias="col_headers",
                                description="Comma-separated translated column headers"),
):
    """Export quotes as XLSX with frozen header row and auto-filter."""
    db = _get_db(request)
    try:
        project = _check_project(db, project_id)
        ids = _parse_quote_ids(quote_ids)
        quotes = extract_quotes_for_export(db, project_id, ids, anonymise=anonymise)

        if not quotes:
            raise HTTPException(status_code=404, detail="No quotes match the filter")

        col_headers = _parse_headers(headers)

        # Lazy import — openpyxl is heavy and only needed for XLSX
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = excel_sheet_name(project.name)

        # Header row (bold)
        bold = Font(bold=True)
        for col_idx, header in enumerate(col_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold

        # Data rows
        for row_idx, q in enumerate(quotes, start=2):
            for col_idx, value in enumerate(_quote_to_row(q), start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter on all columns
        last_col_letter = chr(ord("A") + len(col_headers) - 1)
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(quotes) + 1}"

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(col_headers, start=1):
            max_width = len(header)
            for q in quotes[:100]:  # Sample first 100 rows
                row = _quote_to_row(q)
                val = row[col_idx - 1]
                max_width = max(max_width, min(len(val), 60))
            col_letter = chr(ord("A") + col_idx - 1)
            ws.column_dimensions[col_letter].width = max_width + 2

        # Write to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{safe_filename(project.name)}-quotes.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )
    finally:
        db.close()
