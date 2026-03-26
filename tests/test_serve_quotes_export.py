"""Tests for quotes export endpoints (CSV + XLSX)."""

from __future__ import annotations

import csv
import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from bristlenose.server.app import create_app
from tests.conftest import AuthTestClient

_FIXTURE_DIR = Path(__file__).parent / "fixtures" / "smoke-test" / "input"

# Smoke-test quotes:
# q-p1-10 (Dashboard / confusion)
# q-p1-26 (Dashboard / frustration)
# q-p1-46 (Search / delight)
# q-p1-66 (Onboarding gaps / frustration)


@pytest.fixture()
def client() -> TestClient:
    """Create a test client with imported smoke-test data."""
    app = create_app(project_dir=_FIXTURE_DIR, dev=True, db_url="sqlite://")
    return AuthTestClient(app)


@pytest.fixture()
def unauth_client() -> TestClient:
    """Create a test client WITHOUT auth headers."""
    app = create_app(project_dir=_FIXTURE_DIR, dev=True, db_url="sqlite://")
    return TestClient(app)


# ---------------------------------------------------------------------------
# CSV endpoint
# ---------------------------------------------------------------------------


class TestCsvExport:
    def test_returns_200(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.csv")
        assert resp.status_code == 200

    def test_content_type(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.csv")
        ct = resp.headers.get("content-type", "")
        assert "text/csv" in ct

    def test_content_disposition(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.csv")
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".csv" in cd

    def test_has_utf8_bom(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.csv")
        assert resp.text.startswith("\ufeff")

    def test_has_header_row(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.csv")
        # Strip BOM then parse
        text = resp.text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        assert len(header) == 11

    def test_has_data_rows(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.csv")
        text = resp.text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        # 1 header + 4 data rows
        assert len(rows) == 5

    def test_quote_ids_filter(self, client: TestClient) -> None:
        resp = client.get(
            "/api/projects/1/export/quotes.csv?quote_ids=q-p1-10"
        )
        text = resp.text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert len(rows) == 2  # 1 header + 1 data row

    def test_anonymise_blanks_name(self, client: TestClient) -> None:
        resp = client.get(
            "/api/projects/1/export/quotes.csv?anonymise=true"
        )
        text = resp.text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        name_col = header.index("Participant name")
        for row in reader:
            assert row[name_col] == ""

    def test_csv_formula_injection_defence(self, client: TestClient) -> None:
        """Cell values starting with = + - @ get tab-prefixed."""
        resp = client.get("/api/projects/1/export/quotes.csv")
        text = resp.text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(text))
        next(reader)  # skip header
        for row in reader:
            for cell in row:
                if cell and cell[0] in "=+@":
                    assert cell.startswith("\t")

    def test_404_nonexistent_project(self, client: TestClient) -> None:
        resp = client.get("/api/projects/999/export/quotes.csv")
        assert resp.status_code == 404

    def test_custom_headers(self, client: TestClient) -> None:
        custom = ",".join(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"])
        resp = client.get(f"/api/projects/1/export/quotes.csv?col_headers={custom}")
        text = resp.text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        assert header == ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]


# ---------------------------------------------------------------------------
# XLSX endpoint
# ---------------------------------------------------------------------------


class TestXlsxExport:
    def test_returns_200(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.xlsx")
        assert resp.status_code == 200

    def test_content_type(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.xlsx")
        ct = resp.headers.get("content-type", "")
        assert "spreadsheetml" in ct

    def test_content_disposition(self, client: TestClient) -> None:
        resp = client.get("/api/projects/1/export/quotes.xlsx")
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert ".xlsx" in cd

    def test_valid_xlsx(self, client: TestClient) -> None:
        """Returned bytes should be a valid openpyxl workbook."""
        from openpyxl import load_workbook

        resp = client.get("/api/projects/1/export/quotes.xlsx")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 1 header + 4 data rows
        assert ws.max_row == 5
        assert ws.max_column == 11

    def test_frozen_header(self, client: TestClient) -> None:
        from openpyxl import load_workbook

        resp = client.get("/api/projects/1/export/quotes.xlsx")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_auto_filter(self, client: TestClient) -> None:
        from openpyxl import load_workbook

        resp = client.get("/api/projects/1/export/quotes.xlsx")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.auto_filter.ref is not None

    def test_sheet_name_truncated(self, client: TestClient) -> None:
        """Sheet name should be at most 31 chars."""
        from openpyxl import load_workbook

        resp = client.get("/api/projects/1/export/quotes.xlsx")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert len(ws.title) <= 31

    def test_quote_ids_filter(self, client: TestClient) -> None:
        from openpyxl import load_workbook

        resp = client.get(
            "/api/projects/1/export/quotes.xlsx?quote_ids=q-p1-10,q-p1-26"
        )
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows

    def test_404_nonexistent_project(self, client: TestClient) -> None:
        resp = client.get("/api/projects/999/export/quotes.xlsx")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------


class TestAuthRequired:
    def test_csv_requires_auth(self, unauth_client: TestClient) -> None:
        resp = unauth_client.get("/api/projects/1/export/quotes.csv")
        assert resp.status_code == 401

    def test_xlsx_requires_auth(self, unauth_client: TestClient) -> None:
        resp = unauth_client.get("/api/projects/1/export/quotes.xlsx")
        assert resp.status_code == 401
